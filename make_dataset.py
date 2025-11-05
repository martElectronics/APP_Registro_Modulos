# =======================================================
# Requirements
# =======================================================
import sys
import os
import re
import json
import shutil
import pandas   as pd
import tkinter  as tk

from datetime           import datetime
from openpyxl           import Workbook, load_workbook
from openpyxl.styles    import Font, Border, Side, Alignment, PatternFill
from openpyxl.utils     import get_column_letter
from tkinter            import simpledialog


# =======================================================
# RDirecciones de memoria (EJECUTABLE)
# =======================================================
if getattr(sys, 'frozen', False):
    base_path = os.path.dirname(sys.executable)
else:
    base_path = os.path.dirname(os.path.abspath(__file__))

config_path = os.path.join(base_path, "Config.json")

with open(config_path, "r", encoding="utf-8") as f:
    config = json.load(f)

# =======================================================
# Configuración de rutas
# =======================================================
path_data       = os.path.join(base_path, config["Path"]["Data"])
path_procesed   = os.path.join(base_path, config["Path"]["Data_proccesed"])
file_excel      = os.path.join(base_path, path_procesed, ".backup", "Battery Data.xlsx")

os.makedirs(path_data, exist_ok=True)
os.makedirs(path_procesed, exist_ok=True)
os.makedirs(os.path.join(base_path, path_procesed, ".backup"), exist_ok=True)

Modulos     = config["Modulos"]
Temp_max    = config["Limites"]["Temp max"]
Temp_min    = config["Limites"]["Temp min"]
V_max       = config["Limites"]["V max"]
V_min       = config["Limites"]["V min"]

# =======================================================
# Funciones Auxiliares
# =======================================================

def date_file(archivo):

    patron_completo = r'(\d{2})[-_](\d{2})[-_](\d{4})(?:[-_](\d{1,2})-(\d{1,2}))?'
    match = re.search(patron_completo, archivo)
    if match:

        try:
            dia, mes, año, hora, minuto = match.groups()
            hora = int(hora) if hora else 0
            minuto = int(minuto) if minuto else 0
            fecha = datetime(int(año), int(mes), int(dia), hora, minuto)

            return fecha.strftime("%d-%m-%Y %H-%M")
        except ValueError:

            pass
    return datetime.now().strftime("%Y-%m-%d %H-%M")

def parse_fecha(nombre_hoja):
    try:
        return datetime.strptime(nombre_hoja.split("_")[0], "%Y-%m-%d")
    except ValueError:
        return None

def obtener_color_verde(valor, min_val, max_val):
    if max_val == min_val:
        ratio = 0.5
    else:
        ratio = (valor - min_val) / (max_val - min_val)
        ratio = max(0.0, min(1.0, ratio))

    r_base, g_base, b_base = 0, 176, 80
    r = int(255 - (255 - r_base) * ratio)
    g = int(255 - (255 - g_base) * ratio)
    b = int(255 - (255 - b_base) * ratio)

    return PatternFill(start_color=f"{r:02X}{g:02X}{b:02X}",
                       end_color=f"{r:02X}{g:02X}{b:02X}",
                       fill_type="solid")

def obtener_color_rojo(valor, min_val, max_val):
    if max_val == min_val:
        ratio = 0.5
    else:
        ratio = (valor - min_val) / (max_val - min_val)
        ratio = max(0.0, min(1.0, ratio))

    r_base, g_base, b_base = 255, 0, 0
    r = int(255 - (255 - r_base) * ratio)
    g = int(255 - (255 - g_base) * ratio)
    b = int(255 - (255 - b_base) * ratio)

    return PatternFill(start_color=f"{r:02X}{g:02X}{b:02X}",
                       end_color=f"{r:02X}{g:02X}{b:02X}",
                       fill_type="solid")

# =======================================================
# Lectura y Procesamiento de Archivos
# =======================================================
archivos_csv = [f for f in os.listdir(path_data) if f.endswith(".csv")]
if not archivos_csv:
    sys.exit()

if os.path.exists(file_excel):
    wb = load_workbook(file_excel)
else:
    wb = Workbook()
    wb.remove(wb.active)

# =======================================================
# Procesamiento de datos 
# =======================================================
for file in archivos_csv:
    file_path = os.path.join(path_data, file)
    df = pd.read_csv(file_path, sep=";", header=0, index_col=0, decimal='.')

    df.rename(index=lambda x: f"M{Modulos[x]:02d}", inplace=True)
    df = df.sort_index(key=lambda x: x.str[1:].astype(int))

    
    df_clean = df.copy()

    # Identificamoslos valores (Tensiones y Temperaturas)
    for col in df_clean.columns:
        df_clean[col] = pd.to_numeric(df_clean[col], errors='coerce')

        # Valores de Tension 
        if 'V' in col:
            df_clean[col] = df_clean[col].apply(lambda x: x if (isinstance(x, (int, float)) and V_min <= x <= V_max) else "ERROR")

        # Valores de Temperatura 
        if 'T' in col:
            df_clean[col] = df_clean[col].apply(lambda x: x if (isinstance(x, (int, float)) and Temp_min <= x <= Temp_max) else "ERROR")

    date = date_file(file)
    safe_date = date.replace(":", "-")

    # =======================================================
    # Ventana de Solicitud
    # =======================================================
    root = tk.Tk()
    root.withdraw()

    # Comentario para la hoja procesada 
    comentario = simpledialog.askstring("Comentario", f"Introduce comentario para {file}:")
    comentario = comentario or ""

    # Comentario para la exlusión  
    modulos_excluir_input = simpledialog.askstring(
        "Módulos a excluir",
        f"Introduce los módulos a excluir para {file} (números separados por comas, ejemplo: 3,7,12):"
    )
    root.destroy()


    # Procesamos los datos de los modunos NO excluidos 
    modulos_excluir = []

    if modulos_excluir_input:
        numeros = [n.strip() for n in modulos_excluir_input.replace(" ", "").split(",") if n]
        for n in numeros:
            if n.isdigit():
                modulos_excluir.append(f"M{int(n):02d}")
            elif n.upper().startswith("M") and n[1:].isdigit():
                modulos_excluir.append(f"M{int(n[1:]):02d}")

    nuevo_nombre = safe_date
    contador = 1
    while nuevo_nombre in wb.sheetnames:
        nuevo_nombre = f"{safe_date}_{contador}"
        contador += 1

    ws = wb.create_sheet(title=nuevo_nombre, index=0)
    desplazamiento_filas = 5

    # =======================================================
    # Cabecera
    # =======================================================
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=12)
    ws.cell(row=1, column=1, value=f"Registro de Módulos: {ws.title}").font = Font(size=16, bold=True)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=12)
    ws.cell(row=2, column=1, value=f"Comentario: {comentario}").font = Font(size=12, italic=True)
    ws.cell(row=2, column=1).alignment = Alignment(horizontal="center")

    df_filtrado = df_clean.drop(index=[m for m in modulos_excluir if m in df_clean.index], errors="ignore")

    # =======================================================
    # Función para crear tablas (Voltajes o Temperaturas)
    # =======================================================

    def crear_bloque(df_data, tipo, fila_inicio, color_func):
        cols = [c for c in df_data.columns if tipo in c]
        if not cols:
            return fila_inicio

        color_label = "verde" if tipo == "V" else "rojo"
        titulo      = f"Tabla principal ({'Voltajes' if tipo=='V' else 'Temperaturas'})"
        
        ws.merge_cells(start_row=fila_inicio, start_column=1, end_row=fila_inicio, end_column=len(cols) + 1)
        ws.cell(row=fila_inicio, column=1, value=titulo).font = Font(size=14, bold=True)
        ws.cell(row=fila_inicio, column=1).alignment = Alignment(horizontal="center")

        fila_headers = fila_inicio + 1
        ws.cell(row=fila_headers, column=1, value="Módulo").font = Font(bold=True)
        for c_idx, col_name in enumerate(cols, start=2):
            ws.cell(row=fila_headers, column=c_idx, value=col_name).font = Font(bold=True)
            ws.cell(row=fila_headers, column=c_idx).alignment = Alignment(horizontal="center")

        for r_idx, (idx, row) in enumerate(df_data.iterrows(), start=fila_headers + 1):
            ws.cell(row=r_idx, column=1, value=idx)

            for c_idx, col_name in enumerate(cols, start=2):
                val = row[col_name]
                ws.cell(row=r_idx, column=c_idx, value=val)

            # Procesamos los datos por fila 
            valores_validos = [v for v in row[cols] if isinstance(v, (int, float))]
            if valores_validos:
                vmin, vmax = min(valores_validos), max(valores_validos)
                for c_idx, col_name in enumerate(cols, start=2):
                    val = row[col_name]
                    if isinstance(val, (int, float)):
                        ws.cell(row=r_idx, column=c_idx).fill = color_func(val, vmin, vmax)

        # TABLA DE ESTADISTICAS (POR FILA)
        df_num = df_data[cols].apply(pd.to_numeric, errors='coerce') 
        df_stats = pd.DataFrame({
            "Min"           : df_num.min(axis=1),
            "Max"           : df_num.max(axis=1),
            "Media"         : df_num.mean(axis=1).round(2),
            "Mediana"       : df_num.median(axis=1).round(2),
            "Moda"          : df_num.mode(axis=1).iloc[:, 0].round(2),
            "Desv_Estandar" : df_num.std(axis=1).round(2)
        })

        col_stats = len(cols) + 3
        ws.cell(row=fila_headers, column=col_stats, value=f"Módulo").font = Font(bold=True)
        for c_idx, col_name in enumerate(df_stats.columns, start=col_stats + 1):
            ws.cell(row=fila_headers, column=c_idx, value=col_name).font = Font(bold=True)
            ws.cell(row=fila_headers, column=c_idx).alignment = Alignment(horizontal="center")

        for r_idx, (modulo, fila) in enumerate(df_stats.iterrows(), start=fila_headers + 1):
            ws.cell(row=r_idx, column=col_stats, value=modulo)
            for c_idx, valor in enumerate(fila, start=col_stats + 1):
                ws.cell(row=r_idx, column=c_idx, value=valor)
                ws.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal="center")

        # bloque global
        fila_global = fila_headers + len(df_data) + 3
        ws.merge_cells(start_row=fila_global, start_column=1, end_row=fila_global, end_column=len(cols) + 1)
        ws.cell(row=fila_global, column=1, value=f"Mapa Global {'Voltajes' if tipo=='V' else 'Temperaturas'}").font = Font(bold=True, size=14)
        ws.cell(row=fila_global, column=1).alignment = Alignment(horizontal="center")

        fila_headers_g = fila_global + 1
        ws.cell(row=fila_headers_g, column=1, value="Módulo").font = Font(bold=True)
        for c_idx, col_name in enumerate(cols, start=2):
            ws.cell(row=fila_headers_g, column=c_idx, value=col_name).font = Font(bold=True)
            ws.cell(row=fila_headers_g, column=c_idx).alignment = Alignment(horizontal="center")

        for r_idx, (idx, row) in enumerate(df_data.iterrows(), start=fila_headers_g + 1):
            ws.cell(row=r_idx, column=1, value=idx)
            for c_idx, col_name in enumerate(cols, start=2):
                val = row[col_name]
                ws.cell(row=r_idx, column=c_idx, value=val)


        valores_globales = df_num.stack().dropna().astype(float)
        if not valores_globales.empty:
            vmin, vmax = valores_globales.min(), valores_globales.max()
            for r in range(fila_headers_g + 1, fila_headers_g + 1 + len(df_data)):
                for c in range(2, len(cols) + 2):
                    val = ws.cell(row=r, column=c).value
                    if isinstance(val, (int, float)):
                        ws.cell(row=r, column=c).fill = color_func(val, vmin, vmax)

            # estadísticas globales
            fila_stats_g = fila_headers_g + len(df_data) + 2
            ws.cell(row=fila_stats_g, column=1, value=f"Estadísticas Globales {'Voltaje' if tipo=='V' else 'Temperatura'}").font = Font(bold=True, size=12)
            ws.merge_cells(start_row=fila_stats_g, start_column=1, end_row=fila_stats_g, end_column=len(df_stats.columns) + 1)

            fila_stats_header = fila_stats_g + 1
            for c_idx, col_name in enumerate(df_stats.columns, start=2):
                ws.cell(row=fila_stats_header, column=c_idx, value=col_name).font = Font(bold=True)
                ws.cell(row=fila_stats_header, column=c_idx).alignment = Alignment(horizontal="center")

            # TABLA DE ESTADISTICAS (POR FILA)
            fila_stats_valor = fila_stats_header + 1
            stats_global = {
                "Min"           : vmin,
                "Max"           : vmax,
                "Media"         : valores_globales.mean().round(2),
                "Mediana"       : valores_globales.median().round(2),
                "Moda"          : valores_globales.mode().iloc[0].round(2),
                "Desv_Estandar" : valores_globales.std().round(2)
            }
            for c_idx, key in enumerate(df_stats.columns, start=2):
                ws.cell(row=fila_stats_valor, column=c_idx, value=stats_global[key])

            return fila_stats_valor + 3
        else:
            return fila_headers_g + len(df_data) + 2

    # =======================================================
    # Bloques completos
    # =======================================================
    fila_actual = desplazamiento_filas
    fila_actual = crear_bloque(df_filtrado, "V", fila_actual, obtener_color_verde)
    fila_actual = crear_bloque(df_filtrado, "T", fila_actual + 2, obtener_color_rojo)

    for col_idx in range(1, 40):
        ws.column_dimensions[get_column_letter(col_idx)].width = 14

# =======================================================
# Aplicar bordes negros a todas las celdas con contenido
# =======================================================
border_black = Border(
    left=Side(border_style="thin", color="000000"),
    right=Side(border_style="thin", color="000000"),
    top=Side(border_style="thin", color="000000"),
    bottom=Side(border_style="thin", color="000000")
)

for ws in wb.worksheets:
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None and str(cell.value).strip() != "":
                cell.border = border_black


# =======================================================
# Guardado final
# =======================================================
hojas_con_fecha = [(sh, parse_fecha(sh)) for sh in wb.sheetnames if parse_fecha(sh)]
hojas_con_fecha.sort(key=lambda x: x[1], reverse=True)
hojas_ordenadas = [h[0] for h in hojas_con_fecha] + [h for h in wb.sheetnames if not parse_fecha(h)]
for idx, nombre in enumerate(hojas_ordenadas):
    wb._sheets[idx] = wb[nombre]

wb.save(file_excel)
wb.close()
shutil.copy2(file_excel, base_path)
print("✅ Procesamiento completado correctamente.")